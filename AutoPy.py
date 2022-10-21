import openpyxl

inv_file = openpyxl.load_workbook("C:\\Users\\hmada\\OneDrive\\Desktop\\PyPro\\inventory.xlsx")
product_list = inv_file['Sheet1']

products_per_supplier = {}
total_value_per_supplier = {}
products_uder_10_inv = {}

# We will go each row in the product_list
for product_row in range(2, product_list.max_row + 1):
    # We are trying to get the col 4
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_num = product_list.cell(product_row, 1).value
    inventort_price = product_list.cell(product_row, 5)

    # We are getting the company name with number of product
    if supplier_name in products_per_supplier:
       current_num_products =  products_per_supplier.get(supplier_name)
       products_per_supplier[supplier_name] = current_num_products + 1
    else:
        print("adding new supplier")
        products_per_supplier[supplier_name] = 1

    # Calculation total value of inventory per supplier
    if supplier_name in total_value_per_supplier:
       current_totla_value = total_value_per_supplier.get(supplier_name)
       total_value_per_supplier[supplier_name] = current_totla_value + inventory * price
    else:
        total_value_per_supplier[supplier_name] = inventory * price

    # Printing products that has inventory less 10
    if inventory < 10:
        products_uder_10_inv[product_num] = inventory

    # add value for total inventory price
    inventort_price.value = inventory * price

print(products_uder_10_inv)
print(products_per_supplier)
print(total_value_per_supplier)

# Saving the changes in new file
inv_file.save("C:\\Users\\hmada\\OneDrive\\Desktop\\PyPro\\inventory_with_total_value.xlsx")
